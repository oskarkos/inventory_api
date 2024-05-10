import json
from datetime import datetime, timedelta, date

from django.db.models.functions.datetime import TruncYear, TruncDay, TruncHour, TruncMinute, TruncSecond
from django.db.models.functions.text import Upper
from openpyxl.styles import Font, Alignment
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter

from rest_framework.viewsets import ModelViewSet
from rest_framework.views import APIView

from app_control.models import DianResolution, PaymentTerminal, Provider, Customer
from inventory_api.excel_manager import apply_styles_to_cells

from .serializers import (
    Inventory, InventorySerializer, InventoryGroupSerializer, InventoryGroup,
    Invoice, InvoiceSerializer, InventoryWithSumSerializer,
    InvoiceItem, DianSerializer, PaymentTerminalSerializer, ProviderSerializer, UserWithAmountSerializer,
    CustomerSerializer
)
from rest_framework.response import Response
from rest_framework import status
from inventory_api.custom_methods import IsAuthenticatedCustom
from inventory_api.utils import CustomPagination, get_query, create_terminals_report, create_dollars_report, \
    create_cash_report, create_inventory_report, create_product_sales_report, create_invoices_report
from django.db.models import Count, Sum, F, Q, Value, CharField, Func
from django.db.models.functions import Coalesce, TruncMonth
from user_control.models import CustomUser
import csv
import codecs
from django.http import HttpResponse
from openpyxl import Workbook


class InventoryView(ModelViewSet):
    http_method_names = ('get', 'put', 'delete', 'post')
    queryset = Inventory.objects.select_related("group", "created_by")
    serializer_class = InventorySerializer
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination

    def get_queryset(self):
        if self.request.method.lower() != "get":
            return self.queryset

        data = self.request.query_params.dict()
        data.pop("page", None)
        keyword = data.pop("keyword", None)

        results = self.queryset.filter(**data)

        if keyword:
            search_fields = (
                "code", "created_by__fullname", "created_by__email",
                "group__name", "name"
            )
            query = get_query(keyword, search_fields)
            return results.filter(query)

        return results

    def create(self, request, *args, **kwargs):
        request.data.update({"created_by_id": request.user.id})
        return super().create(request, *args, **kwargs)

    def update(self, request, pk=None):
        inventory = Inventory.objects.filter(pk=pk).first()
        serializer = self.serializer_class(inventory, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None):
        inventory = Inventory.objects.filter(pk=pk).first()
        inventory.delete()
        return Response({"message": "Inventory deleted successfully"}, status=status.HTTP_200_OK)

    def toggle_active(self, request, pk=None):
        inventory = Inventory.objects.filter(pk=pk).first()
        if inventory is None:
            return Response({'error': 'Inventory not found'}, status=status.HTTP_404_NOT_FOUND)

        inventory.active = not inventory.active
        inventory.save()
        serializer = self.serializer_class(inventory)
        return Response(serializer.data)


class ProviderView(ModelViewSet):
    http_method_names = ('get', 'put', 'delete', 'post')
    queryset = Provider.objects.select_related("created_by")
    serializer_class = ProviderSerializer
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination

    def get_queryset(self):
        data = self.request.query_params.dict()
        keyword = data.pop("keyword", None)
        data.pop("page", None)
        results = self.queryset.filter(**data)

        if keyword:
            search_fields = (
                "nit", "created_by__fullname", "created_by__email", "name"
            )
            query = get_query(keyword, search_fields)
            results = results.filter(query)

        return results.order_by('id')

    def create(self, request, *args, **kwargs):
        request.data.update({"created_by_id": request.user.id})
        return super().create(request, *args, **kwargs)

    def update(self, request, pk):
        provider = Provider.objects.filter(pk=pk).first()
        serializer = self.serializer_class(provider, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk):
        provider = Provider.objects.filter(pk=pk).first()
        provider.delete()
        return Response({"message": "Provider deleted successfully"}, status=status.HTTP_200_OK)

    def toggle_active(self, request, pk=None):
        provider = Provider.objects.filter(pk=pk).first()
        if provider is None:
            return Response({'error': 'Provider not found'}, status=status.HTTP_404_NOT_FOUND)

        provider.active = not provider.active
        provider.save()
        serializer = self.serializer_class(provider)
        return Response(serializer.data)


class CustomerView(ModelViewSet):
    http_method_names = ('get', 'post', 'put', 'delete')
    queryset = Customer.objects.select_related(
        "created_by")
    serializer_class = CustomerSerializer
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination

    def get_queryset(self):
        if self.request.method.lower() != "get":
            return self.queryset
        data = self.request.query_params.dict()
        data.pop("page", None)
        keyword = data.pop("keyword", None)

        results = self.queryset.filter(**data).order_by('id')

        if keyword:
            search_fields = (
                "created_by__fullname", "created_by__email", "document_id", "name"
            )
            query = get_query(keyword, search_fields)
            results = results.filter(query)

        return results.order_by('id')

    def create(self, request, *args, **kwargs):
        request.data.update({"created_by_id": request.user.id})
        return super().create(request, *args, **kwargs)

    def destroy(self, request, pk):
        customer = Customer.objects.filter(pk=pk).first()
        customer.delete()
        return Response({"message": "Customer deleted successfully"}, status=status.HTTP_200_OK)


class InventoryGroupView(ModelViewSet):
    http_method_names = ('get', 'put', 'delete', 'post')

    queryset = InventoryGroup.objects.select_related(
        "belongs_to", "created_by").prefetch_related("inventories")
    serializer_class = InventoryGroupSerializer
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination

    def get_queryset(self):
        if self.request.method.lower() != "get":
            return self.queryset
        data = self.request.query_params.dict()
        data.pop("page", None)

        keyword = data.pop("keyword", None)
        results = self.queryset.filter(**data).order_by('id')

        if keyword:
            search_fields = (
                "created_by__fullname", "created_by__email", "name"
            )
            query = get_query(keyword, search_fields)
            results = results.filter(query)

        return results.annotate(
            total_items=Count('inventories')
        )

    def create(self, request, *args, **kwargs):
        request.data.update({"created_by_id": request.user.id})
        return super().create(request, *args, **kwargs)

    def update(self, request, pk=None):
        inventory_group = InventoryGroup.objects.filter(pk=pk).first()
        serializer = self.serializer_class(inventory_group, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None):
        inventory_group = InventoryGroup.objects.filter(pk=pk).first()
        inventory_group.delete()
        return Response({"message": "Inventory Group deleted successfully"}, status=status.HTTP_200_OK)


class PaymentTerminalView(ModelViewSet):
    http_method_names = ('get', 'post', 'put', 'delete')
    queryset = PaymentTerminal.objects.select_related("created_by")
    serializer_class = PaymentTerminalSerializer
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination

    def get_queryset(self):
        if self.request.method.lower() != "get":
            return self.queryset
        data = self.request.query_params.dict()
        data.pop("page", None)

        keyword = data.pop("keyword", None)
        results = self.queryset.filter(**data)

        if keyword:
            search_fields = (
                "created_by__fullname", "created_by__email", "name", "account_code"
            )
            query = get_query(keyword, search_fields)
            results = results.filter(query)

        return results.order_by('id')

    def create(self, request, *args, **kwargs):
        request.data.update({"created_by_id": request.user.id})
        return super().create(request, *args, **kwargs)

    def update(self, request, pk=None):
        terminal = PaymentTerminal.objects.filter(pk=pk).first()
        serializer = self.serializer_class(terminal, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None):
        terminal = PaymentTerminal.objects.filter(pk=pk).first()
        terminal.delete()
        return Response({"message": "Payment Terminal deleted successfully"}, status=status.HTTP_200_OK)

    def toggle_active(self, request, pk=None):
        terminal = PaymentTerminal.objects.filter(pk=pk).first()
        if terminal is None:
            return Response({'error': 'Payment Terminal not found'}, status=status.HTTP_404_NOT_FOUND)

        terminal.active = not terminal.active
        terminal.save()
        serializer = self.serializer_class(terminal)
        return Response(serializer.data)


class InvoiceView(ModelViewSet):
    http_method_names = ('get', 'post', 'put', 'delete')
    queryset = Invoice.objects.select_related(
        "created_by", "sale_by", "payment_terminal", "dian_resolution").prefetch_related("invoice_items")
    serializer_class = InvoiceSerializer
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination

    def get_queryset(self):
        if self.request.method.lower() != "get":
            return self.queryset

        data = self.request.query_params.dict()
        data.pop("page", None)

        keyword = data.pop("keyword", None)
        results = self.queryset.filter(**data)

        if keyword:
            search_fields = (
                "created_by__fullname", "created_by__email", "invoice_number", "dian_resolution__document_number",
            )
            query = get_query(keyword, search_fields)
            results = results.filter(query)

        return results

    def create(self, request, *args, **kwargs):
        dian_resolution = DianResolution.objects.filter(active=True).first()
        if not dian_resolution:
            raise Exception("You need to have an active dian resolution to create invoices")

        try:
            if not request.data.get("sale_by_id"):
                request.data.update({"sale_by_id": request.user.id})

            request.data.update({"created_by_id": request.user.id})

            new_current_number = dian_resolution.current_number + 1
            dian_resolution_document_number = dian_resolution.id
            dian_resolution.current_number = new_current_number
            dian_resolution.save()

            request.data.update(
                {"dian_resolution_id": dian_resolution_document_number, "invoice_number": new_current_number})
            return super().create(request, *args, **kwargs)
        except Exception as e:
            dian_resolution.current_number -= 1
            dian_resolution.save()
            raise e

    def update(self, request, pk=None):
        invoice = Invoice.objects.filter(pk=pk).first()
        serializer = self.serializer_class(invoice, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None):
        invoice = Invoice.objects.filter(pk=pk).first()
        invoice.delete()
        return Response({"message": "Invoice deleted successfully"}, status=status.HTTP_200_OK)


class UpdateInvoiceView(APIView):
    def patch(self, request, invoice_number):
        try:
            invoice = Invoice.objects.get(invoice_number=invoice_number)
        except Invoice.DoesNotExist:
            return Response({"error": "Invoice not found"}, status=status.HTTP_404_NOT_FOUND)

        if invoice.is_override:
            return Response({"error": "Invoice already overrided"}, status=status.HTTP_400_BAD_REQUEST)

        # Actualizar el estado is_override de la factura a True
        invoice.is_override = True

        # Restaurar la cantidad de elementos en el inventario para los InvoiceItems correspondientes
        for item in invoice.invoice_items.all():
            inventory_item = item.item
            inventory_item.total_in_shops += item.quantity
            inventory_item.save()

        # Guardar los cambios en la base de datos
        invoice.save()

        return Response({"message": "Invoice updated successfully"}, status=status.HTTP_200_OK)


class SummaryView(ModelViewSet):
    http_method_names = ('get',)
    permission_classes = (IsAuthenticatedCustom,)
    queryset = InventoryView.queryset

    def list(self, request, *args, **kwargs):
        total_inventory = InventoryView.queryset.filter(
            total_in_storage__gt=0
        ).count()
        total_group = InventoryGroupView.queryset.count()
        total_users = CustomUser.objects.filter(is_superuser=False).count()

        return Response({
            "total_inventory": total_inventory,
            "total_group": total_group,
            "total_users": total_users
        })


class InvoicePainterView(ModelViewSet):
    http_method_names = ('get',)
    permission_classes = (IsAuthenticatedCustom,)

    def list(self, request, *args, **kwargs):
        invoice_number = request.query_params.get("invoice_number", None)
        if not id:
            return Response({"error": "You need to provide invoice id"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            invoice = Invoice.objects.select_related(
                "payment_terminal", "created_by"
            ).filter(invoice_number=invoice_number).first()

            if not invoice:
                return Response({"error": "Invoice not found"}, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response(InvoiceSerializer(invoice).data)


class SalePerformance(ModelViewSet):
    http_method_names = ('get',)
    permission_classes = (IsAuthenticatedCustom,)
    queryset = InventoryView.queryset

    def list(self, request, *args, **kwargs):
        query_data = request.query_params.dict()
        total = query_data.get('total', None)
        query = self.queryset

        if not total:
            start_date = query_data.get("start_date", None)
            end_date = query_data.get("end_date", None)

            if start_date:
                query = query.filter(
                    inventory_invoices__created_at__range=[
                        start_date, end_date]
                )
        items = query.filter(inventory_invoices__invoice__is_override=False).annotate(
            sum_of_item=Coalesce(
                Sum("inventory_invoices__quantity"), 0
            )
        ).order_by('-sum_of_item')[0:10]

        response_data = InventoryWithSumSerializer(items, many=True).data
        return Response(response_data)


class SalesByUsersView(ModelViewSet):
    http_method_names = ('get',)
    permission_classes = (IsAuthenticatedCustom,)

    def list(self, request, *args, **kwargs):
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)

        if not start_date or not end_date:
            sales_by_user = (
                Invoice.objects.select_related("InvoiceItems", "sale_by")
                .all()
                .filter(is_override=False)
                .values_list(
                    "sale_by__fullname"
                )
                .annotate(
                    total_invoice=Sum("invoice_items__amount"),
                )
            )
        else:
            sales_by_user = (
                Invoice.objects.select_related("InvoiceItems", "sale_by")
                .all()
                .filter(is_override=False)
                .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
                .values(
                    "sale_by__id",
                    "sale_by__fullname"
                )
                .annotate(
                    total_invoice=Sum("invoice_items__amount"),
                )
            )

        return Response(sales_by_user)


class PurchaseView(ModelViewSet):
    http_method_names = ('get',)
    permission_classes = (IsAuthenticatedCustom,)
    queryset = InvoiceView.queryset

    def list(self, request, *args, **kwargs):
        query_data = request.query_params.dict()
        total = query_data.get('total', None)
        query = InvoiceItem.objects.select_related("invoice", "item")
        start_date = query_data.get("start_date", None)
        end_date = query_data.get("end_date", None)

        if not total and start_date:
            query = query.filter(
                Q(create_at__range=[start_date, end_date]) & Q(
                    invoice__is_override=False)
            )
        else:
            query = query.exclude(invoice__is_override=True)

        results = query.aggregate(
            amount_total=Sum(F('amount')),
            total=Sum('quantity'),
            amount_total_usd=Sum(F('usd_amount'),
                                 filter=Q(invoice__is_dollar=True, invoice__is_override=False))
        )

        selling_price = results.get("amount_total", 0.0)
        count = results.get("total", 0)
        price_dolar = results.get("amount_total_usd", 0.0)

        response_data = {
            "count": count,
        }

        if selling_price is not None:
            response_data["selling_price"] = "{:.2f}".format(selling_price)

        if price_dolar is not None:
            response_data["price_dolar"] = "{:.2f}".format(price_dolar)

        return Response(response_data)


class InventoryCSVLoaderView(ModelViewSet):
    http_method_names = ('post',)
    queryset = InventoryView.queryset
    permission_classes = (IsAuthenticatedCustom,)
    serializer_class = InventorySerializer

    def create(self, request, *args, **kwargs):
        try:
            data = request.FILES['data']
        except Exception as e:
            raise Exception("you need to provide inventory CSV data")

        inventory_items = []

        try:
            csv_reader = csv.reader(
                codecs.iterdecode(data, 'utf-8'), delimiter=',')
            next(csv_reader)
            for row in csv_reader:
                if not row[0]:
                    continue
                inventory_items.append({
                    "group_id": int(row[0]),
                    "code": str(row[1]),
                    "name": str(row[2]),
                    "photo": str(row[3]),
                    "total_in_storage": int(row[4]),
                    "total_in_shops": int(row[5]),
                    "selling_price": float(row[6]),
                    "buying_price": float(row[7]),
                    "usd_price": float(row[8]),
                    "provider_id": int(row[9]),
                    "created_by_id": request.user.id
                })
        except csv.Error as e:
            raise Exception(e)

        if not inventory_items:
            raise Exception("CSV file cannot be empty")

        data_validation = self.serializer_class(
            data=inventory_items, many=True)
        data_validation.is_valid(raise_exception=True)
        data_validation.save()

        return Response({
            "success": "Inventory items added successfully"
        })


class DianResolutionView(ModelViewSet):
    http_method_names = ('get', 'put', 'delete', 'post')
    queryset = DianResolution.objects.all()
    serializer_class = DianSerializer
    permission_classes = (IsAuthenticatedCustom,)
    pagination_class = CustomPagination

    def get_queryset(self):
        if self.request.method.lower() != 'get':
            return self.queryset

        data = self.request.query_params.dict()
        data.pop("page", None)
        keyword = data.pop("keyword", None)

        results = self.queryset.filter(**data)

        if keyword:
            search_fields = (
                "created_by", "document_number", "current_number"
            )
            query = get_query(keyword, search_fields)
            return results.filter(query)

        return results

    def create(self, request, *args, **kwargs):
        request.data.update({"created_by_id": request.user.id})

        if DianResolution.objects.all().filter(active=True).exists():
            raise Exception("You can't have more than one active dian resolution, "
                            "please deactivate the current one first")

        return super().create(request, *args, **kwargs)

    def update(self, request, pk=None):
        dian_res = DianResolution.objects.filter(pk=pk).first()
        serializer = self.serializer_class(dian_res, data=request.data)

        if request.data.get("active", True) is True:
            if DianResolution.objects.all().filter(active=True).exists():
                raise Exception("You can't have more than one active dian resolution, "
                                "please deactivate the current one first")

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None):
        dian_res = DianResolution.objects.filter(pk=pk).first()
        dian_res.delete()
        return Response({"message": "Dian Resolution deleted successfully"}, status=status.HTTP_200_OK)

    def toggle_active(self, request, pk=None):
        resolution = DianResolution.objects.filter(pk=pk).first()
        if resolution is None:
            return Response({'error': 'Dian Resolution not found'}, status=status.HTTP_404_NOT_FOUND)

        if DianResolution.objects.all().filter(active=True).exists() and resolution.active is False:
            raise Exception("You can't have more than one active dian resolution, "
                            "please deactivate the current one first")

        resolution.active = not resolution.active
        resolution.save()
        serializer = self.serializer_class(resolution)
        return Response(serializer.data)


class ReportExporter(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request):
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_ventas_{start_date}_al_{end_date}.xlsx"'

        if not start_date or not end_date:
            return Response({"error": "You need to provide start_date and end_date"})

        # cast start_date and end_date to datetime
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

        wb = Workbook()
        ws = wb.active
        ws.title = "REPORTE DIARIO"
        ws.column_dimensions[get_column_letter(2)].width = 29

        terminals_report_data = (
            Invoice.objects.select_related("PaymentMethods", "payment_terminal", "created_by")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(payment_methods__name__in=["debitCard", "creditCard"])
            .values_list("payment_terminal__name", "sale_by__fullname")
            .annotate(
                quantity=Count("id"),
                total=Sum("payment_methods__paid_amount")
            )
        )

        last_row_cards = create_terminals_report(ws, terminals_report_data, start_date, end_date)

        dollar_report_data = (
            Invoice.objects.select_related("InvoiceItems", "created_by")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(is_dollar=True)
            .values_list("sale_by__fullname")
            .annotate(
                quantity=Sum("invoice_items__usd_amount")
            )
        )

        last_row_dollars = create_dollars_report(ws, dollar_report_data, last_row_cards, start_date, end_date)

        cash_report_data = (
            Invoice.objects.select_related("InvoiceItems", "created_by")
            .all()
            .filter(is_override=False)
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .values_list("sale_by__fullname")
            .annotate(
                quantity=Sum("invoice_items__amount")
            )
        )

        dollar_report_data_in_pesos = (
            Invoice.objects.select_related("InvoiceItems", "created_by")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(is_dollar=True)
            .values_list("sale_by__fullname")
            .annotate(
                quantity=Sum("invoice_items__amount")
            )
        )

        cards_report_data = (
            Invoice.objects.select_related("PaymentMethods", "created_by")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(payment_methods__name__in=["debitCard", "creditCard"])
            .values_list("sale_by__fullname")
            .annotate(
                total=Sum("payment_methods__paid_amount")
            )
        )

        last_row, last_column = create_cash_report(ws, last_row_dollars, last_row_cards,
                                                   cash_report_data, dollar_report_data_in_pesos, cards_report_data,
                                                   start_date, end_date
                                                   )

        # center all text in the cells from A1 to the last cell
        apply_styles_to_cells(1, 1, last_column, last_row, ws, alignment=Alignment(horizontal="center"))

        wb.save(response)
        return response


class InventoriesReportExporter(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_inventarios.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "REPORTE DE INVENTARIOS"

        inventories_report_data = (
            Inventory.objects.select_related("group")
            .all()
            .annotate(
                upper_group_name=Upper("group__belongs_to__name"),
                upper_group_subname=Upper("group__name"),
                upper_name=Upper("name"),
                total_price_in_shops=F('total_in_shops') * F('buying_price'),
                total_price_in_storage=F('total_in_storage') * F('buying_price'),
                total_selling_price_in_shops=F('total_in_shops') * F('selling_price'),
                total_selling_price_in_storage=F('total_in_storage') * F('selling_price'),
                units=Value("COP", output_field=CharField())
            )
            .values_list("upper_group_name", "upper_group_subname", "code", "upper_name", "total_in_storage",
                         "total_in_shops", "buying_price", "selling_price",
                         "total_price_in_shops", "total_price_in_storage", "total_selling_price_in_shops",
                         "total_selling_price_in_storage", "units"
                         )
        )

        create_inventory_report(ws, inventories_report_data)

        wb.save(response)
        return response


class ItemsReportExporter(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request):
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_ventas_x_producto_{start_date}_{end_date}.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "REPORTE DE VENTAS POR PRODUCTO"

        if not start_date or not end_date:
            return Response({"error": "You need to provide start_date and end_date"})

        report_data = (
            Invoice.objects.select_related("InvoiceItems")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(invoice_items__is_gift=False)
            .values_list("invoice_items__item_code", "invoice_items__item_name")
            .annotate(
                quantity=Sum("invoice_items__quantity"),
            )
        )

        report_data_nulled = (
            Invoice.objects.select_related("InvoiceItems")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=True)
            .values_list("invoice_items__item_code", "invoice_items__item_name")
            .annotate(
                quantity=Sum("invoice_items__quantity"),
            )
        )

        report_data_gifts = (
            Invoice.objects.select_related("InvoiceItems")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .filter(invoice_items__is_gift=True)
            .values_list("invoice_items__item_code", "invoice_items__item_name")
            .annotate(
                quantity=Sum("invoice_items__quantity"),
            )
        )

        create_product_sales_report(ws, report_data, report_data_nulled, report_data_gifts, start_date, end_date)

        wb.save(response)
        return response


class InvoicesReportExporter(APIView):
    http_method_names = ('post',)
    permission_classes = (IsAuthenticatedCustom,)

    def post(self, request):
        start_date = request.data.get("start_date", None)
        end_date = request.data.get("end_date", None)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_facturas_{start_date}_{end_date}.xlsx"'

        if not start_date or not end_date:
            return Response({"error": "You need to provide start_date and end_date"})

        wb = Workbook()
        ws = wb.active
        ws.title = "REPORTE DE FACTURACION"

        inventories_report_data = (
            Invoice.objects.select_related("payment_terminal", "InvoiceItems", "created_by")
            .all()
            .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .filter(is_override=False)
            .annotate(
                total_invoice=Sum("invoice_items__amount"),
            )
            .values_list(
                "created_at__date", "sale_by__fullname", "invoice_number", "dian_resolution__document_number",
                "payment_terminal__name", "total_invoice",
                "customer__document_id", "customer__name", "customer__email", "customer__phone",
                "customer__address"
            )
        )

        create_invoices_report(ws, inventories_report_data)

        wb.save(response)
        return response
